from django.shortcuts import render,redirect,get_object_or_404
from django.db.models import *
from checkout.models import *
from orders.models import *
from .models import *
from django.contrib.auth import authenticate,login,logout
from user.views import Home
from django.contrib.auth.models import User
from django.utils.dateparse import parse_date
from django.db.models import Sum
from datetime import date, timedelta
from django.http import HttpResponse
from django.core.paginator import Paginator
from reportlab.lib.pagesizes import A4 # type: ignore
from reportlab.lib import colors # type: ignore
from reportlab.lib.units import inch # type: ignore
from reportlab.lib.styles import getSampleStyleSheet # type: ignore
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph,Image, Spacer # type: ignore
from io import BytesIO
from reportlab.pdfgen import canvas # type: ignore
from openpyxl import Workbook # type: ignore
from reportlab.lib import colors # type: ignore
import openpyxl # type: ignore
from openpyxl import Workbook # type: ignore
from openpyxl.styles import Font, Alignment # type: ignore
from django.utils.dateparse import parse_date
from functools import wraps



def active_admin(func):
    @wraps(func)
    def _wrapped_view(req, *args, **kwargs):
        if req.user.is_staff==True and req.user.is_authenticated:
            
            return func(req, *args, **kwargs)
        else:
            print('fbvdj')
            return redirect(admin_login)
    return _wrapped_view


def admin_login(req):
    if req.method=='POST':
        username = req.POST['username']
        password = req.POST['password']
        user = authenticate(username = username ,password = password)
        if user:
            login(req,user)
            return redirect('admin_dashboard')
    return render(req,'login.html')


@active_admin
def admin_dashboard(req):
    
            sales = Orders.objects.all().exclude(Q(confirm = False) | Q(status = 'canceled')).order_by('-id')
            user_count = User.objects.all().count()
            order_count = Orders.objects.all().count()
            product_count = Product.objects.all().count()
            date_range = req.GET.get('date_range')
            start_date = req.GET.get('start_date')
            end_date = req.GET.get('end_date')

            if date_range == 'daily':
                sales = sales.filter(order_date=date.today()).exclude(Q(confirm = False) | Q(status = 'canceled')).order_by('-id')
            elif date_range == 'weekly':
                sales = sales.filter(order_date__range=(date.today()-timedelta(days=7), date.today())).exclude(Q(confirm = False) | Q(status = 'canceled')).order_by('-id')      
            elif date_range == 'monthly':
                sales = sales.filter(order_date__month=date.today().month).exclude(Q(confirm = False) | Q(status = 'canceled')).order_by('-id')
            elif date_range == 'yearly':
                sales = sales.filter(order_date__year=date.today().year).exclude(Q(confirm = False) | Q(status = 'canceled')).order_by('-id')
            elif date_range == 'custom' and start_date and end_date:
                start_date = parse_date(start_date)
                end_date = parse_date(end_date)
                sales = sales.filter(order_date__range=(start_date, end_date)).exclude(Q(confirm = False) | Q(status = 'canceled')).order_by('-id')

            # Calculate overall metrics
            overall_sales_count = sales.count()
            overall_order_amount = sales.aggregate(Sum('grand_total'))['grand_total__sum'] or 0
            overall_discount = sales.aggregate(Sum('discount'))['discount__sum'] or 0
            overall_total = sales.aggregate(Sum('total'))['total__sum'] or 0
            
             

            context = {
                'user_count' : user_count,
                'order_count' : order_count,
                'product_count' : product_count,
                'page_obj': sales,
                'overall_sales_count': overall_sales_count,
                'overall_order_amount': overall_order_amount,
                'date_range' : date_range,
                'overall_discount': overall_discount,
                'date_ranges' : date_range,
                'start_dates' : start_date,
                'end_dates' : end_date

            }
            return render(req,'admin_dashboard.html',context)

@active_admin
def user_details(req):
    obj = User.objects.exclude(is_staff=True).order_by('-id')
    context = {
        'obj':obj
    }
    return render(req,'user.html',context)




def toggle_status(req, id):
    user = get_object_or_404(User, id=id)
    user.is_active = not user.is_active
    user.save()
    return redirect(user_details)


def admin_logout(req):
    logout(req)
    return redirect(Home)



def generate_pdf_report(request, date_range, start_date, end_date):
    sales = Orders.objects.all().exclude(Q(confirm=False) | Q(status='canceled')).order_by('-id')
    note = ''
    if date_range == 'daily':
        sales = sales.filter(order_date=date.today())
        note = 'Daily Report'
    elif date_range == 'weekly':
        sales = sales.filter(order_date__range=(date.today() - timedelta(days=7), date.today()))
        note = 'Weekly Report'
    elif date_range == 'monthly':
        sales = sales.filter(order_date__month=date.today().month)
        note = 'Monthly Report'
    elif date_range == 'yearly':
        sales = sales.filter(order_date__year=date.today().year)
        note = 'Yearly Report'
    elif date_range == 'custom' and start_date and end_date:
        start_date = parse_date(start_date)
        end_date = parse_date(end_date)
        sales = sales.filter(order_date__range=(start_date, end_date))
        note = f'Sale between {start_date} -- {end_date}'

    # Calculate overall metrics
    overall_sales_count = sales.count()
    overall_order_amount = sales.aggregate(Sum('grand_total'))['grand_total__sum'] or 0
    overall_discount = sales.aggregate(Sum('discount'))['discount__sum'] or 0
    overall_total = sales.aggregate(Sum('total'))['total__sum'] or 0
    success_count = sales.filter(status = 'success').count()
    pending_count = sales.filter(status = 'pending').count()
    today = timezone.now().date()
    
    buffer = BytesIO()
    pdf = SimpleDocTemplate(buffer, pagesize=A4)

    image_path = 'assets\images\soundsphere-removebg-preview.png'
    img = Image(image_path, width=100, height=50) 
    elements = []
    data = [[img, ''],  # Image in the first cell, empty space in the second cell
            ['', '']]   # Adding an empty row below to push the table content down

    table = Table(data, colWidths=[400, 100])  # Adjust colWidths as needed

    # Build the PDF
    elements = [table, Spacer(1, 20)]
    styles = getSampleStyleSheet()
    title = Paragraph("Sales Report", styles['Title'])
    elements.append(title)
    elements.append(Spacer(1, 12))

    # Overall metrics
    metrics = [
        note,
        f"Delivered Orders: {success_count}",
        f"Pending Orders: {pending_count}",
        f"Sales Count: {overall_sales_count}",
        f"Total Amount: Rs {overall_total:.2f}",
        f"Discount Given: Rs {overall_discount:.2f}",
        f"Order Amount: Rs {overall_order_amount:.2f}",
        f"Date : {today}"
    ]
    
    for metric in metrics:
        elements.append(Paragraph(metric, styles['BodyText']))
        elements.append(Spacer(1, 12))

    # Spacer before the table
    elements.append(Spacer(1, 24))

    # Example data for the table
    data = [
        ['Date', 'Order_id', 'Customer', 'Amount', 'Discount', 'Final Amount', 'Payment Method', 'Status'],
    ]
    for sale in sales:
        data.append([
            sale.order_date,
            sale.id,
            sale.user_id.username,
            sale.total,
            sale.discount,
            sale.grand_total,
            sale.payment,
            sale.status,
        ])

    # Create the table
    table = Table(data)

    # Apply style to the table
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
    ]))

    elements.append(table)
    pdf.build(elements)

    pdf_value = buffer.getvalue()
    buffer.close()

    response = HttpResponse(pdf_value, content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="report.pdf"'

    return response



def generate_excel_report(request, date_range, start_date, end_date):
    sales = Orders.objects.all().exclude(Q(confirm=False) | Q(status='canceled')).order_by('-id')
    note = ''
    if date_range == 'daily':
        sales = sales.filter(order_date=date.today())
        note = 'Daily Report'
    elif date_range == 'weekly':
        note = 'Weekly Report'
        sales = sales.filter(order_date__range=(date.today() - timedelta(days=7), date.today()))
    elif date_range == 'monthly':
        note = 'monthly Report'
        sales = sales.filter(order_date__month=date.today().month)
    elif date_range == 'yearly':
        sales = sales.filter(order_date__year=date.today().year)
        note = 'Yearly Report'
    elif date_range == 'custom' and start_date and end_date:
        start_date = parse_date(start_date)
        end_date = parse_date(end_date)
        sales = sales.filter(order_date__range=(start_date, end_date))
        note = f'Sales between {start_date} -- {end_date}'

    # Calculate overall metrics
    overall_sales_count = sales.count()
    overall_order_amount = sales.aggregate(Sum('grand_total'))['grand_total__sum'] or 0
    overall_discount = sales.aggregate(Sum('discount'))['discount__sum'] or 0
    overall_total = sales.aggregate(Sum('total'))['total__sum'] or 0
    success_count = sales.filter(status='success').count()
    pending_count = sales.filter(status='pending').count()
    today = date.today()

    # Create a new Workbook and select the active worksheet
    wb = Workbook()
    ws = wb.active
    ws.title = "Sales Report"

    # Add title and metrics
    ws.merge_cells('A1:H1')
    title = 'Sales Report'
    ws['A1'] = title
    ws['A1'].font = Font(size=14, bold=True)
    ws['A1'].alignment = Alignment(horizontal='center')

    metrics = [
        f'Note: {note}',
        f'Delivered Orders: {success_count}',
        f'Pending Orders: {pending_count}',
        f'Sales Count: {overall_sales_count}',
        f'Total Amount: Rs {overall_total:.2f}',
        f'Discount Given: Rs {overall_discount:.2f}',
        f'Order Amount: Rs {overall_order_amount:.2f}',
        f'Date: {today}'
    ]

    row = 3  # Start adding metrics from the third row
    for metric in metrics:
        ws[f'A{row}'] = metric
        row += 1

    # Add headers for the table
    headers = ['Date', 'Order ID', 'Customer', 'Amount', 'Discount', 'Final Amount', 'Payment Method', 'Status']
    ws.append(headers)

    # Style the header row
    for cell in ws[4]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')
    
    # Add sales data to the worksheet
    for sale in sales:
        ws.append([
            sale.order_date.strftime('%Y-%m-%d'),
            sale.id,
            sale.user_id.username,
            sale.total,
            sale.discount,
            sale.grand_total,
            sale.payment,
            sale.status,
        ])

    # Set column widths for better readability
    column_widths = [15, 10, 20, 15, 10, 15, 20, 15]
    for i, width in enumerate(column_widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = width

    # Save the workbook to a BytesIO object
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="sales_report.xlsx"'

    wb.save(response)

    return response





